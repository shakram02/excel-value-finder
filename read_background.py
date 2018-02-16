from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.styles.colors import RGB
from openpyxl.worksheet import Worksheet
from argparse import ArgumentParser

optionParser = ArgumentParser()

optionParser.add_argument("--needle", help="Smaller set of items to be searched for in haystack",
                          default="responses.xlsx")

optionParser.add_argument("--column", help="Column name contianing needle values",
                          default="B")


# optionParser.add_argument("haystack", help="Bigger file to search for items")


def main():
    args = optionParser.parse_args()

    wb = load_workbook(filename=args.needle)
    form_responses: Worksheet = wb['Form Responses 1']
    name_set = set()

    out_of_scope_colors = ["FF434343", "FF999999"]
    out_of_scope = 0

    for item in form_responses[args.column]:
        cell: Cell = item
        if cell.value is None:
            continue

        # print(cell.value)
        name_set.update(cell.value)

        if cell.fill.bgColor.rgb in out_of_scope_colors:
            # print("Out of scope")
            out_of_scope += 1

    print("{}/{} rejected".format(out_of_scope, form_responses.max_row))
    print(form_responses['B18'].fill.bgColor.rgb)


if __name__ == "__main__":
    main()
