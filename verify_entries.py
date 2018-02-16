from argparse import ArgumentParser

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import Worksheet

optionParser = ArgumentParser()

optionParser.add_argument("--needle", help="Smaller set of items to be searched for in haystack",
                          default="responses.xlsx")

optionParser.add_argument("--column", help="Column name contianing needle values",
                          default="B")


# optionParser.add_argument("haystack", help="Bigger file to search for items")

def filter_column(ws_column, is_accepted, is_unique=None, on_duplicate=None):
    """
    Pass on all form entries, add cell values if the cell meets acceptance criteria

    Checking for duplicates won't work without providing both ``is_unique`` and ``on_duplicate``


    :param ws_column: Column of the worksheet (idc about field header, caller removes it)
    :param is_accepted: Function pointer to test acceptance
    :param is_unique: Function pointer to check uniqueness
    :param on_duplicate: Callback function on finding duplicates
    :return:
    """
    accepted = []
    rejected = []

    for item in ws_column:
        cell: Cell = item

        if cell.value is None:
            continue

        value = cell.value.strip()

        if is_accepted(cell):
            accepted.append(value)
            if is_unique is not None and on_duplicate is not None and not is_unique(cell):
                on_duplicate(cell)
        else:
            rejected.append(value)

    return accepted, rejected


def extract_valid_applicants(wb_path, sheet_name, column, reject_colors):
    wb = load_workbook(filename=wb_path)

    form_responses: Worksheet = wb[sheet_name]
    name_set = set()

    # TODO, handle duplicates, create a triple of name, phone, phone or name, mail, phone or whateva
    def on_dup(c):
        print("Duplicate:", c.value)

    def is_accepted(c):
        return c.fill.bgColor.rgb not in reject_colors

    def is_unique(c):
        if c.value in name_set:
            return False

        name_set.add(c.value)
        return True

    names, rejected = filter_column(form_responses[column][1:], is_accepted, is_unique, on_dup)
    print("{}/{} didn't pass application phase".format(len(names), form_responses.max_row))

    from pprint import pprint
    pprint(name_set)
    print("\n\n\n\n\n\n\n\n")
    pprint(names)
    print("Unique:", len(name_set), "Unique:", len(names))
    return names


def get_called_applicants():
    pass


def main():
    # Protons form verification, we want to make sure that we didn't forget to call anyone

    args = optionParser.parse_args()

    # Don't forget to add FF (alpha channel) at the begning of the color
    reject_colors = ["FF434343", "FF999999"]
    sheet_name = 'Form Responses 1'
    applicants = extract_valid_applicants(args.needle, sheet_name, args.column, reject_colors)
    print(len(applicants))


# Main
if __name__ == "__main__":
    main()
